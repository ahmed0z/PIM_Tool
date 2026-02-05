import streamlit as st
import pandas as pd
import pickle
import os
import glob
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl
import tempfile

# Constants
PRESET_DB_PATH = "preset_db.pkl"

def load_preset_db():
    """Load the preset database from file if it exists."""
    if os.path.exists(PRESET_DB_PATH):
        with open(PRESET_DB_PATH, 'rb') as f:
            return pickle.load(f)
    return None

def cleanup_old_preset_files():
    """Remove old preset Excel and PKL files (except preset_db.pkl)."""
    app_dir = os.path.dirname(os.path.abspath(__file__))
    patterns = ['*.pkl', '*.xlsx', '*.xlsm', '*.xltx', '*.xltm']
    removed = []
    for pattern in patterns:
        for filepath in glob.glob(os.path.join(app_dir, pattern)):
            filename = os.path.basename(filepath)
            # Keep only preset_db.pkl, remove all other pkl/excel files
            if filename != 'preset_db.pkl':
                try:
                    os.remove(filepath)
                    removed.append(filename)
                except Exception:
                    pass
    return removed

def save_preset_db(df):
    """Save the preset database to file and cleanup old files."""
    # Cleanup old preset files first
    cleanup_old_preset_files()
    # Save new preset
    with open(PRESET_DB_PATH, 'wb') as f:
        pickle.dump(df, f)

def run_full_process(pim_file_bytes, part_data_file_bytes, progress_bar, status_text):
    """Run the full PIM processing workflow."""
    try:
        # Load preset DB
        preset_df = load_preset_db()
        if preset_df is None:
            status_text.error("No preset database found. Please upload one in the Settings page.")
            return None, None

        # Save uploaded files to temp files for openpyxl
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_pim:
            tmp_pim.write(pim_file_bytes)
            pim_file = tmp_pim.name
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_part:
            tmp_part.write(part_data_file_bytes)
            part_data_file = tmp_part.name

        # Step 1: Load PIM file
        status_text.info("Loading PIM file...")
        progress_bar.progress(10)
        wb = load_workbook(pim_file)
        ws = wb.worksheets[0]
        max_row = ws.max_row

        # --- Step 1: Move columns N and O to R and S ---
        cols_to_move = [14, 15]  # N, O
        moved_data = []
        for col_idx in cols_to_move:
            col_data = []
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                col_data.append({
                    'value': cell.value,
                    'style': cell._style,
                    'number_format': cell.number_format,
                    'has_formula': cell.data_type == 'f',
                    'formula': cell.value if cell.data_type == 'f' else None
                })
            moved_data.append(col_data)

        ws.delete_cols(15)
        ws.delete_cols(14)
        ws.insert_cols(18, amount=2)

        for i, col_data in enumerate(moved_data):
            col_idx = 18 + i
            for row_idx, cell_info in enumerate(col_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = cell_info['value']
                cell._style = cell_info['style']
                cell.number_format = cell_info['number_format']
                if cell_info['has_formula']:
                    cell.value = cell_info['formula']

        # --- Step 2: Copy columns C, D, E, F and insert at N, O, P, Q ---
        copy_cols = [3, 4, 5, 6]
        insert_at = 14

        copied_data = []
        for col_idx in copy_cols:
            col_data = []
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                col_data.append({
                    'value': cell.value,
                    'style': cell._style,
                    'number_format': cell.number_format,
                    'has_formula': cell.data_type == 'f',
                    'formula': cell.value if cell.data_type == 'f' else None
                })
            copied_data.append(col_data)

        ws.insert_cols(insert_at, amount=4)

        for i, col_data in enumerate(copied_data):
            col_idx = insert_at + i
            for row_idx, cell_info in enumerate(col_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = cell_info['value']
                cell._style = cell_info['style']
                cell.number_format = cell_info['number_format']
                if cell_info['has_formula']:
                    cell.value = cell_info['formula']

        progress_bar.progress(30)

        # --- Step 3: Delete column P and insert empty column with header 'XXXXX' ---
        ws.delete_cols(16)
        ws.insert_cols(16)
        ws.cell(row=1, column=16).value = 'XXXXX'

        # --- Step 4: Insert 5 empty columns after Q ---
        ws.insert_cols(18, amount=5)

        # --- Step 5: Rename and format headers for R, S, T ---
        header_map = {18: 'S', 19: 'N', 20: 'D'}
        header_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        header_font = Font(bold=True, color='000000')
        header_align = Alignment(horizontal='center', vertical='center')

        for col_idx, new_name in header_map.items():
            cell = ws.cell(row=1, column=col_idx)
            cell.value = new_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # --- Step 6: Format column V header and columns N, P headers ---
        v_col = 22
        v_cell = ws.cell(row=1, column=v_col)
        v_cell.value = 'Datasheet'
        v_cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        v_cell.font = Font(bold=True, color='000000')
        v_cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in [14, 16]:
            cell = ws.cell(row=1, column=col)
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            cell.font = Font(bold=True, color='9C0006')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        progress_bar.progress(40)

        # --- Step 7: Concatenate columns for matching rows only ---
        filter_keywords = ["new", "check updates", "check value"]
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=8).value
            if cell_value:
                cell_str = str(cell_value).lower()
                if any(keyword in cell_str for keyword in filter_keywords):
                    n_val = ws.cell(row=row_idx, column=14).value or ""
                    o_val = ws.cell(row=row_idx, column=15).value or ""
                    ws.cell(row=row_idx, column=16).value = f"{n_val}{o_val}"
                    l_val = ws.cell(row=row_idx, column=12).value or ""
                    m_val = ws.cell(row=row_idx, column=13).value or ""
                    ws.cell(row=row_idx, column=21).value = f"{l_val}{m_val}"

        # --- Step 8: Process part data file ---
        status_text.info("Processing part data file...")
        wb2 = load_workbook(part_data_file)
        ws2 = wb2.worksheets[0]
        max_row2 = ws2.max_row
        ws2.insert_cols(5)
        ws2.cell(row=1, column=5).value = ''
        for row_idx in range(2, max_row2 + 1):
            c_val = ws2.cell(row=row_idx, column=3).value or ""
            d_val = ws2.cell(row=row_idx, column=4).value or ""
            ws2.cell(row=row_idx, column=5).value = f"{c_val}{d_val}"
        wb2.save(part_data_file)
        progress_bar.progress(50)

        # --- Step 9: Lookup from part data file ---
        status_text.info("Processing part data lookup...")
        part_wb = load_workbook(part_data_file, data_only=True)
        part_ws = part_wb.worksheets[0]
        lookup_col = 5
        q_col = 17
        s_col = 19
        part_lookup = {}
        for row in range(2, part_ws.max_row + 1):
            key = part_ws.cell(row=row, column=lookup_col).value
            q_val = part_ws.cell(row=row, column=q_col).value
            s_val = part_ws.cell(row=row, column=s_col).value
            if key is not None:
                part_lookup[str(key)] = (q_val, s_val)

        for row_idx in range(2, ws.max_row + 1):
            h_val = ws.cell(row=row_idx, column=8).value
            if h_val:
                h_str = str(h_val).lower()
                if any(keyword in h_str for keyword in filter_keywords):
                    lookup_key = ws.cell(row=row_idx, column=21).value
                    if lookup_key is not None and str(lookup_key) in part_lookup:
                        q_val, s_val = part_lookup[str(lookup_key)]
                        if q_val and 'nod' in str(q_val).lower():
                            ws.cell(row=row_idx, column=22).value = s_val
                        else:
                            ws.cell(row=row_idx, column=22).value = q_val
        progress_bar.progress(70)

        # --- Step 10: COUNTIF-style counts ---
        filtered_row_indices = []
        filtered_n = []
        filtered_p = []
        filtered_v = []
        for row_idx in range(2, ws.max_row + 1):
            h_val = ws.cell(row=row_idx, column=8).value
            if h_val:
                h_str = str(h_val).lower()
                if any(keyword in h_str for keyword in filter_keywords):
                    filtered_row_indices.append(row_idx)
                    filtered_n.append(ws.cell(row=row_idx, column=14).value)
                    filtered_p.append(ws.cell(row=row_idx, column=16).value)
                    filtered_v.append(ws.cell(row=row_idx, column=22).value)

        for i, row_idx in enumerate(filtered_row_indices):
            n_val = filtered_n[i]
            p_val = filtered_p[i]
            v_val = filtered_v[i]
            count_n = sum(1 for val in filtered_n if val == n_val and val not in [None, ""])
            count_p = sum(1 for val in filtered_p if val == p_val and val not in [None, ""])
            count_v = sum(1 for val in filtered_v if val == v_val and val not in [None, ""])
            ws.cell(row=row_idx, column=18).value = count_n
            ws.cell(row=row_idx, column=19).value = count_p
            ws.cell(row=row_idx, column=20).value = count_v

        progress_bar.progress(85)

        # --- Step 11: Lookup from preset source ---
        status_text.info("Processing preset lookup and generating output...")
        lookup_values = []
        for row_idx in filtered_row_indices:
            val = ws.cell(row=row_idx, column=16).value
            if val is not None and val != "":
                lookup_values.append(str(val))

        matched_rows = preset_df[preset_df.iloc[:, 4].astype(str).isin(lookup_values)]

        # --- Step 12: Final formatting of PIM file ---
        # Remove column U (21) as it's no longer needed
        ws.delete_cols(21)
        
        # Apply auto filter to all columns
        max_col = ws.max_column
        max_row = ws.max_row
        filter_range = f"A1:{get_column_letter(max_col)}1"
        ws.auto_filter.ref = filter_range
        
        # Add borders to all cells with text and set column widths to fit headers
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths based on header text and add borders to cells with content
        for col_idx in range(1, max_col + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            header_value = header_cell.value
            if header_value:
                # Set column width to fit header text (with extra padding for filter icon)
                col_width = len(str(header_value)) + 5
                ws.column_dimensions[get_column_letter(col_idx)].width = col_width
            
            # Add borders to all cells with text in this column
            for row_idx in range(1, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip() != '':
                    cell.border = thin_border

        # Save PIM file to bytes
        pim_output = BytesIO()
        wb.save(pim_output)
        pim_output.seek(0)

        # Create preset output file
        preset_output = None
        if not matched_rows.empty:
            preset_output = BytesIO()
            matched_rows.to_excel(preset_output, index=False)
            preset_output.seek(0)
            
            # Apply formatting
            preset_output_formatted = BytesIO()
            wb_out = openpyxl.load_workbook(preset_output)
            ws_out = wb_out.active
            
            green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            pink_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
            black_font = Font(color="000000")
            red_font = Font(color="9C0006")

            for col_letter in ['D', 'E', 'F']:
                if ws_out[f'{col_letter}1'].value is not None:
                    if col_letter == 'D':
                        ws_out[f'{col_letter}1'].fill = green_fill
                    elif col_letter == 'E':
                        ws_out[f'{col_letter}1'].fill = pink_fill
                    elif col_letter == 'F':
                        ws_out[f'{col_letter}1'].fill = blue_fill

            for row in ws_out.iter_rows(min_row=1, max_row=ws_out.max_row):
                for cell in row:
                    if cell.column_letter == 'E':
                        cell.font = red_font
                    else:
                        cell.font = black_font

            ws_out.column_dimensions['D'].width = 37
            ws_out.column_dimensions['E'].width = 80
            ws_out.column_dimensions['F'].width = 95
            
            wb_out.save(preset_output_formatted)
            preset_output_formatted.seek(0)
            preset_output = preset_output_formatted

        progress_bar.progress(100)
        status_text.success("All steps completed successfully!")

        # Cleanup temp files
        os.unlink(pim_file)
        os.unlink(part_data_file)

        return pim_output, preset_output

    except Exception as e:
        status_text.error(f"Error: {str(e)}")
        return None, None


def main_page():
    """Main processing page."""
    st.title("🔧 PIM Format Automation Tool")
    st.markdown("---")

    # Initialize session state for results
    if 'pim_output' not in st.session_state:
        st.session_state.pim_output = None
    if 'preset_output' not in st.session_state:
        st.session_state.preset_output = None
    if 'process_complete' not in st.session_state:
        st.session_state.process_complete = False

    # Check if preset DB exists
    preset_exists = os.path.exists(PRESET_DB_PATH)
    if not preset_exists:
        st.warning("⚠️ No preset database found. Please go to **Settings** page to upload one.")
    else:
        st.success("✅ Preset database loaded")

    st.subheader("Upload Files")
    
    col1, col2 = st.columns(2)
    
    with col1:
        pim_file = st.file_uploader(
            "📄 PIM File", 
            type=['xlsx', 'xlsm', 'xltx', 'xltm'],
            help="Upload your PIM Issue Report Excel file",
            key="pim_file"
        )
    
    with col2:
        part_data_file = st.file_uploader(
            "📄 Part Data File", 
            type=['xlsx', 'xlsm', 'xltx', 'xltm'],
            help="Upload your Part Data Excel file",
            key="part_data_file"
        )

    # Clear results when new files are uploaded
    if pim_file is None or part_data_file is None:
        if st.session_state.process_complete:
            st.session_state.pim_output = None
            st.session_state.preset_output = None
            st.session_state.process_complete = False

    st.markdown("---")

    if st.button("🚀 Run Process", type="primary", disabled=not preset_exists):
        if not pim_file or not part_data_file:
            st.error("Please upload both files before running.")
        else:
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            pim_output, preset_output = run_full_process(
                pim_file.getvalue(),
                part_data_file.getvalue(),
                progress_bar,
                status_text
            )
            
            if pim_output:
                st.session_state.pim_output = pim_output.getvalue()
                st.session_state.preset_output = preset_output.getvalue() if preset_output else None
                st.session_state.process_complete = True

    # Show download buttons if results exist
    if st.session_state.process_complete and st.session_state.pim_output:
        st.markdown("---")
        st.subheader("📥 Download Results")
        
        col1, col2 = st.columns(2)
        
        current_date = datetime.now().strftime("%d_%m_%Y")
        
        with col1:
            st.download_button(
                label="📥 Download Processed PIM File",
                data=st.session_state.pim_output,
                file_name=f"PIM_Processed_{current_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col2:
            if st.session_state.preset_output:
                st.download_button(
                    label="📥 Download DK Preset File",
                    data=st.session_state.preset_output,
                    file_name=f"DK_Preset_{current_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.info("No matching preset records found.")


def settings_page():
    """Settings page for managing preset database."""
    st.title("⚙️ Settings")
    st.markdown("---")

    st.subheader("Preset Database Management")

    # Show current status
    if os.path.exists(PRESET_DB_PATH):
        preset_df = load_preset_db()
        file_stats = os.stat(PRESET_DB_PATH)
        modified_time = datetime.fromtimestamp(file_stats.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        
        st.success(f"✅ Preset database exists")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Rows", f"{len(preset_df):,}")
        with col2:
            st.metric("Columns", len(preset_df.columns))
        with col3:
            st.metric("Last Updated", modified_time)
        
        with st.expander("Preview Database (first 10 rows)"):
            st.dataframe(preset_df.head(10))
    else:
        st.warning("⚠️ No preset database found. Please upload an Excel or PKL file below.")

    st.markdown("---")
    st.subheader("Update Preset Database")
    
    st.markdown("**Option 1: Upload file through browser**")
    uploaded_file = st.file_uploader(
        "📤 Upload Excel or PKL file to create/update preset database",
        type=['xlsx', 'xlsm', 'xltx', 'xltm', 'pkl'],
        help="Excel files will be converted to PKL format. This will replace the existing database."
    )

    if uploaded_file:
        file_ext = uploaded_file.name.split('.')[-1].lower()
        st.info(f"File selected: {uploaded_file.name} ({file_ext.upper()} format)")
        
        if st.button("💾 Save as Preset Database", type="primary"):
            try:
                with st.spinner("Processing..."):
                    if file_ext == 'pkl':
                        # Load directly from PKL
                        df = pickle.load(uploaded_file)
                        st.info("Loaded from PKL file")
                    else:
                        # Load from Excel and convert to PKL
                        df = pd.read_excel(uploaded_file)
                        st.info("Converted Excel to PKL format")
                    
                    save_preset_db(df)
                st.success(f"✅ Database saved successfully! ({len(df):,} rows)")
                st.rerun()
            except Exception as e:
                st.error(f"Error saving database: {str(e)}")

    st.markdown("---")
    st.markdown("**Option 2: Load from local file path (for large files)**")
    st.caption("Place your file in the repo folder and enter the filename below")
    
    local_file = st.text_input("File path (e.g., `preset_source.xlsx` or `preset_source.pkl`)")
    
    if local_file and st.button("📂 Load from Local File", type="secondary"):
        file_path = os.path.join(os.path.dirname(__file__), local_file)
        if os.path.exists(file_path):
            try:
                with st.spinner("Processing..."):
                    if file_path.lower().endswith('.pkl'):
                        with open(file_path, 'rb') as f:
                            df = pickle.load(f)
                        st.info("Loaded from PKL file")
                    else:
                        df = pd.read_excel(file_path)
                        st.info("Converted Excel to PKL format")
                    
                    save_preset_db(df)
                st.success(f"✅ Database saved successfully! ({len(df):,} rows)")
                st.rerun()
            except Exception as e:
                st.error(f"Error loading file: {str(e)}")
        else:
            st.error(f"File not found: {file_path}")

    st.markdown("---")
    
    # Danger zone
    if os.path.exists(PRESET_DB_PATH):
        with st.expander("🗑️ Danger Zone"):
            st.warning("This action cannot be undone!")
            if st.button("Delete Preset Database", type="secondary"):
                os.remove(PRESET_DB_PATH)
                st.success("Database deleted.")
                st.rerun()


# Page navigation
st.set_page_config(
    page_title="PIM Format Tool",
    page_icon="🔧",
    layout="wide"
)

# Sidebar navigation
st.sidebar.title("Navigation")
page = st.sidebar.radio("Go to", ["🏠 Main", "⚙️ Settings"])

if page == "🏠 Main":
    main_page()
else:
    settings_page()
