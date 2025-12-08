from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Alignment
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import pickle
from datetime import datetime
import pandas as pd
import os
import openpyxl
import subprocess

# File path
input_file = 'test of PIM Issue Report_17072025_Final.xlsx'

def run_full_process(pim_file, part_data_file, preset_source_file, status_callback, progress_callback, done_callback):
    try:
        # Step 1: Load PIM file
        status_callback("Loading PIM file...")
        progress_callback(10)
        wb = load_workbook(pim_file)
        ws = wb.worksheets[0]
        max_row = ws.max_row

        # --- Step 1: Move columns N and O to R and S (already implemented) ---
        # Columns: N=14, O=15, R=18, S=19 (1-based indexing)
        # Copy columns N and O data (and styles)
        cols_to_move = [14, 15]  # N, O
        moved_data = []
        for col_idx in cols_to_move:
            col_data = []
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                col_data.append({
                    'value': cell.value,
                    'style': cell._style,
                    'number_format': cell.number_format,
                    'has_formula': cell.data_type == 'f',
                    'formula': cell.value if cell.data_type == 'f' else None
                })
            moved_data.append(col_data)

        # Delete columns N and O (delete O first, then N)
        ws.delete_cols(15)
        ws.delete_cols(14)

        # Insert 2 new columns at R (now index 17 after deletion)
        ws.insert_cols(18, amount=2)

        # Paste moved data into new R and S
        for i, col_data in enumerate(moved_data):
            col_idx = 18 + i  # R and S
            for row_idx, cell_info in enumerate(col_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = cell_info['value']
                cell._style = cell_info['style']
                cell.number_format = cell_info['number_format']
                if cell_info['has_formula']:
                    cell.value = cell_info['formula']

        # --- Step 2: Copy columns C, D, E, F and insert at N, O, P, Q ---
        # C=3, D=4, E=5, F=6; N=14, O=15, P=16, Q=17 (1-based)
        copy_cols = [3, 4, 5, 6]
        insert_at = 14

        # Copy data and styles from C, D, E, F
        copied_data = []
        for col_idx in copy_cols:
            col_data = []
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                col_data.append({
                    'value': cell.value,
                    'style': cell._style,
                    'number_format': cell.number_format,
                    'has_formula': cell.data_type == 'f',
                    'formula': cell.value if cell.data_type == 'f' else None
                })
            copied_data.append(col_data)

        # Insert 4 new columns at N (14)
        ws.insert_cols(insert_at, amount=4)

        # Paste copied data into new N, O, P, Q
        for i, col_data in enumerate(copied_data):
            col_idx = insert_at + i
            for row_idx, cell_info in enumerate(col_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = cell_info['value']
                cell._style = cell_info['style']
                cell.number_format = cell_info['number_format']
                if cell_info['has_formula']:
                    cell.value = cell_info['formula']

        # --- Step 3: Delete column P and insert empty column with header 'XXXXX' ---
        # After previous steps, P is at index 16 (1-based)
        ws.delete_cols(16)
        ws.insert_cols(16)
        ws.cell(row=1, column=16).value = 'XXXXX'

        # --- Step 4: Insert 5 empty columns after Q ---
        # Q is at column 17, so insert at 18
        ws.insert_cols(18, amount=5)

        # --- Step 5: Rename and format headers for R, S, T ---
        # R=18, S=19, T=20 (1-based)
        header_map = {18: 'S', 19: 'N', 20: 'D'}
        header_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        header_font = Font(bold=True, color='000000')
        header_align = Alignment(horizontal='center', vertical='center')

        for col_idx, new_name in header_map.items():
            cell = ws.cell(row=1, column=col_idx)
            cell.value = new_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # --- Step 6: Format column V header and columns N, P headers ---
        # V=22, N=14, P=16 (1-based)
        # Format V
        v_col = 22
        v_cell = ws.cell(row=1, column=v_col)
        v_cell.value = 'Datasheet'
        v_cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        v_cell.font = Font(bold=True, color='000000')
        v_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Format N and P
        for col in [14, 16]:
            cell = ws.cell(row=1, column=col)
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            cell.font = Font(bold=True, color='9C0006')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # --- Step 7: Concatenate columns for matching rows only, do not filter rows ---
        # H=8, N=14, O=15, P=16, L=12, M=13, U=21 (1-based)
        filter_keywords = ["new", "check updates", "check value"]
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=8).value
            if cell_value:
                cell_str = str(cell_value).lower()
                if any(keyword in cell_str for keyword in filter_keywords):
                    n_val = ws.cell(row=row_idx, column=14).value or ""
                    o_val = ws.cell(row=row_idx, column=15).value or ""
                    ws.cell(row=row_idx, column=16).value = f"{n_val}{o_val}"
                    l_val = ws.cell(row=row_idx, column=12).value or ""
                    m_val = ws.cell(row=row_idx, column=13).value or ""
                    ws.cell(row=row_idx, column=21).value = f"{l_val}{m_val}"

        # --- Step 8: Process another file for C&D concatenation into new E ---
        status_callback("Processing part data file...")
        wb2 = load_workbook(part_data_file)
        ws2 = wb2.worksheets[0]
        max_row2 = ws2.max_row
        # Insert new column at E (5)
        ws2.insert_cols(5)
        ws2.cell(row=1, column=5).value = ''  # Optionally set a header
        # Concatenate C & D into new E
        for row_idx in range(2, max_row2 + 1):
            c_val = ws2.cell(row=row_idx, column=3).value or ""
            d_val = ws2.cell(row=row_idx, column=4).value or ""
            ws2.cell(row=row_idx, column=5).value = f"{c_val}{d_val}"
        wb2.save(part_data_file)
        status_callback(f"Processed {part_data_file}: Inserted new column E and concatenated C&D into E.")
        progress_callback(50)

        # --- Step 9: Lookup from part data file and update Datasheet column in PIM file ---
        status_callback("Processing part data lookup...")
        # Load part data into a dict for fast lookup
        part_wb = load_workbook(part_data_file, data_only=True)
        part_ws = part_wb.worksheets[0]
        # Use column E (5) as the lookup key, Q (17) and S (19) as value columns
        lookup_col = 5
        q_col = 17
        s_col = 19
        # Build lookup dict: value in E -> (Q, S)
        part_lookup = {}
        for row in range(2, part_ws.max_row + 1):
            key = part_ws.cell(row=row, column=lookup_col).value
            q_val = part_ws.cell(row=row, column=q_col).value
            s_val = part_ws.cell(row=row, column=s_col).value
            if key is not None:
                part_lookup[str(key)] = (q_val, s_val)
        # Now update the PIM file for filtered rows
        filter_keywords = ["new", "check updates", "check value"]
        for row_idx in range(2, ws.max_row + 1):
            h_val = ws.cell(row=row_idx, column=8).value
            if h_val:
                h_str = str(h_val).lower()
                if any(keyword in h_str for keyword in filter_keywords):
                    lookup_key = ws.cell(row=row_idx, column=21).value  # U
                    if lookup_key is not None and str(lookup_key) in part_lookup:
                        q_val, s_val = part_lookup[str(lookup_key)]
                        if q_val and 'nod' in str(q_val).lower():
                            ws.cell(row=row_idx, column=22).value = s_val  # V
                        else:
                            ws.cell(row=row_idx, column=22).value = q_val  # V
        wb.save(pim_file)
        status_callback(f"Step 9: Performed lookup from part data file (using column E as reference, Q and S by position) and updated Datasheet column in PIM file.")
        progress_callback(70)

        # --- Step 10: COUNTIF-style counts for each filtered row among filtered rows only ---
        filter_keywords = ["new", "check updates", "check value"]
        filtered_row_indices = []
        filtered_n = []
        filtered_p = []
        filtered_v = []
        for row_idx in range(2, ws.max_row + 1):
            h_val = ws.cell(row=row_idx, column=8).value
            if h_val:
                h_str = str(h_val).lower()
                if any(keyword in h_str for keyword in filter_keywords):
                    filtered_row_indices.append(row_idx)
                    filtered_n.append(ws.cell(row=row_idx, column=14).value)
                    filtered_p.append(ws.cell(row=row_idx, column=16).value)
                    filtered_v.append(ws.cell(row=row_idx, column=22).value)

        # For each filtered row, count occurrences of its value in N, P, V among filtered rows
        for i, row_idx in enumerate(filtered_row_indices):
            n_val = filtered_n[i]
            p_val = filtered_p[i]
            v_val = filtered_v[i]
            count_n = sum(1 for val in filtered_n if val == n_val and val not in [None, ""])
            count_p = sum(1 for val in filtered_p if val == p_val and val not in [None, ""])
            count_v = sum(1 for val in filtered_v if val == v_val and val not in [None, ""])
            ws.cell(row=row_idx, column=18).value = count_n  # R
            ws.cell(row=row_idx, column=19).value = count_p  # S
            ws.cell(row=row_idx, column=20).value = count_v  # T

        wb.save(pim_file)
        status_callback(f"Step 10: For each filtered row, counted occurrences of N, P, V values among filtered rows and wrote in R, S, T.")
        progress_callback(85)

        # --- Step 11: Lookup from preset source using filtered P values ---
        status_callback("Processing preset source file and exporting results...")
        if preset_source_file.lower().endswith('.pkl'):
            print(f"Loading database from {preset_source_file}...")
            with open(preset_source_file, 'rb') as f:
                preset_df = pickle.load(f)
        else:
            # Create/load DB from Excel
            print(f"Reading source Excel file: {preset_source_file}")
            preset_df = pd.read_excel(preset_source_file)
            # Save as pkl for future use
            db_path = os.path.splitext(preset_source_file)[0] + '.pkl'
            print(f"Saving database to: {db_path}")
            with open(db_path, 'wb') as f:
                pickle.dump(preset_df, f)
            print("Database created and saved successfully")

        # Get filtered values from column P in PIM file
        lookup_values = []
        for row_idx in filtered_row_indices:
            val = ws.cell(row=row_idx, column=16).value
            if val is not None and val != "":
                lookup_values.append(str(val))
        print(f"Performing lookup for {len(lookup_values)} values...")
        # Perform lookup: match values in column E (5th col, 0-based index 4)
        matched_rows = preset_df[preset_df.iloc[:, 4].astype(str).isin(lookup_values)]

        # Use PIM file's directory for output
        pim_dir = os.path.dirname(pim_file)
        current_date = datetime.now().strftime("%d_%m_%Y")
        output_file = os.path.join(pim_dir, f"DK Preset_{current_date}.xlsx")
        if not matched_rows.empty:
            print("Saving results...")
            matched_rows.to_excel(output_file, index=False)
            # Apply formatting (as in presetlookup3.py)
            try:
                wb_out = openpyxl.load_workbook(output_file)
                ws_out = wb_out.active
                # Define colors
                green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                pink_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
                # Define fonts
                black_font = Font(color="000000")
                red_font = Font(color="9C0006")
                # Apply header formatting (columns D, E, F if present)
                for col_letter in ['D', 'E', 'F']:
                    if ws_out[f'{col_letter}1'].value is not None:
                        if col_letter == 'D':
                            ws_out[f'{col_letter}1'].fill = green_fill
                        elif col_letter == 'E':
                            ws_out[f'{col_letter}1'].fill = pink_fill
                        elif col_letter == 'F':
                            ws_out[f'{col_letter}1'].fill = blue_fill
                # Apply font colors
                for row in ws_out.iter_rows(min_row=1, max_row=ws_out.max_row):
                    for cell in row:
                        if cell.column_letter == 'E':
                            cell.font = red_font
                        else:
                            cell.font = black_font
                # Set column widths
                ws_out.column_dimensions['D'].width = 37
                ws_out.column_dimensions['E'].width = 80
                ws_out.column_dimensions['F'].width = 95
                wb_out.save(output_file)
            except Exception as e:
                print(f"Could not apply formatting: {e}")
            print(f"Found {len(matched_rows)} matching rows")
            print(f"Results saved and formatted: {output_file}")
        else:
            print("No matching records found")
        progress_callback(100)
        status_callback("All steps completed successfully! Results saved.")
        done_callback(os.path.dirname(output_file))

    except Exception as e:
        status_callback(f"Error: {str(e)}")
        progress_callback(0)

def main_gui():
    root = tk.Tk()
    root.title("PIM Format Automation Tool")
    root.geometry("600x450")

    # File path variables
    pim_file = tk.StringVar()
    part_data_file = tk.StringVar()
    preset_source_file = tk.StringVar()
    status_text = tk.StringVar()

    def browse_file(var, filetypes):
        filename = filedialog.askopenfilename(filetypes=filetypes)
        if filename:
            var.set(filename)

    def run_process():
        if not (pim_file.get() and part_data_file.get() and preset_source_file.get()):
            messagebox.showerror("Missing Input", "Please select all required files.")
            return
        status_text.set("Running... Please wait.")
        progress_bar['value'] = 0
        def thread_func():
            run_full_process(
                pim_file.get(),
                part_data_file.get(),
                preset_source_file.get(),
                status_text.set,
                lambda v: root.after(0, progress_bar.config, {'value': v}),
                lambda folder: root.after(0, on_done, folder)
            )
        threading.Thread(target=thread_func, daemon=True).start()

    def on_done(folder):
        messagebox.showinfo("Process Complete", "All steps completed!")
        root.destroy()

    # Layout
    tk.Label(root, text="PIM File:").pack(anchor='w', padx=10, pady=(20,0))
    tk.Entry(root, textvariable=pim_file, width=60).pack(anchor='w', padx=10)
    tk.Button(root, text="Browse", command=lambda: browse_file(pim_file, [("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")])).pack(anchor='w', padx=10, pady=(0,10))

    tk.Label(root, text="Part Data File:").pack(anchor='w', padx=10)
    tk.Entry(root, textvariable=part_data_file, width=60).pack(anchor='w', padx=10)
    tk.Button(root, text="Browse", command=lambda: browse_file(part_data_file, [("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")])).pack(anchor='w', padx=10, pady=(0,10))

    tk.Label(root, text="Source File for Lookup (Excel or .pkl):").pack(anchor='w', padx=10)
    tk.Entry(root, textvariable=preset_source_file, width=60).pack(anchor='w', padx=10)
    tk.Button(root, text="Browse", command=lambda: browse_file(preset_source_file, [("Excel or Pickle files", "*.xlsx *.xlsm *.xltx *.xltm *.pkl")])).pack(anchor='w', padx=10, pady=(0,10))

    progress_bar = ttk.Progressbar(root, orient='horizontal', length=500, mode='determinate')
    progress_bar.pack(pady=20)

    tk.Button(root, text="Run", command=run_process, bg="#00B050", fg="white", font=("Arial", 12, "bold")).pack(pady=10)
    tk.Label(root, textvariable=status_text, fg="blue", wraplength=500, justify='left').pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    main_gui() 